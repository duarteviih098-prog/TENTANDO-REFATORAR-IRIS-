"""Rotas /pagamentos/* e APIs relacionadas."""
import io
import re
from collections import defaultdict
import calendar
from app.os.pdf import _draw_pdf_header, excel_file, table_pdf
from app.shared.cache import clear_view_cache
from app.shared.formatters import br_money, br_now, now_str, parse_br_date, parse_num
from app.shared.months import month_reference_matches_selected, normalize_month_reference
from app.shared.payments import payment_status_is_paid
from app.shared.queries import safe_int_id as _safe_int_id
from app.shared.rows import row_get_value, row_to_dict
from app.storage import backup_company_data

from flask import flash, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.auth.decorators import require_permission
from app.pagamentos.services import (
    build_payment_attachment_items,
    ensure_pagamentos_valid_ids,
    import_pagamentos_excel,
    pagamentos_query_rows,
    pagamentos_totais_from_rows,
    payment_month_or_current,
    save_pagamento,
)
from app.storage import ATTACHMENT_GROUPS, missing_attachment_response, storage_or_local_response, sync_payment_attachments


def _payment_month_or_current(value=''):
    from app.pagamentos.services import payment_month_or_current
    return payment_month_or_current(value)

def query_one(sql, params=()):
    from app.db import query_one as db_query_one
    return db_query_one(sql, params)

def query_all(sql, params=()):
    from app.db import query_all as db_query_all
    return db_query_all(sql, params)

def execute(sql, params=()):
    from app.db import execute as db_execute
    return db_execute(sql, params)

def current_company_id():
    from app.auth import current_company_id as fn
    return fn()

def company_where(table, prefix=' WHERE '):
    from app.auth import company_where as fn
    return fn(table, prefix)

def company_and(table):
    from app.auth import company_and as fn
    return fn(table)

def owned_by_current_company(table, rid):
    from app.auth.tenancy import owned_by_current_company as fn
    return fn(table, rid)

def get_current_user():
    from app.auth import get_current_user as fn
    return fn()

def app_logger():
    from flask import current_app
    return current_app.logger

@require_permission('edit_pagamentos')
def pagamentos_import():
    file = request.files.get('arquivo_excel')
    if not file or not file.filename:
        flash('Selecione um arquivo Excel para importar pagamentos.', 'warning')
        return redirect(url_for('pagamentos'))
    mes_override = request.form.get('mes_importacao', '').strip() or None
    try:
        qtd = import_pagamentos_excel(file, mes_override=mes_override)
        clear_view_cache()
        flash(f'Importação de pagamentos concluída: {qtd} linha(s).', 'success')
    except Exception as exc:
        flash(f'Erro ao importar pagamentos: {exc}', 'danger')
    return redirect(url_for('pagamentos'))





@require_permission('view_pagamentos')
def pagamentos_redirect():
    """Redireciona /pagamentos para o hub."""
    return redirect(url_for('pagamentos_hub'))


@require_permission('view_pagamentos')
def pagamentos():
    """Tela principal de Pagamentos: listagem real com id, filtros e totais coerentes."""
    agora = br_now()
    mes_atual_real = f"{agora.month:02d}/{agora.year:04d}"

    todos = str(request.args.get('todos') or '').strip().lower() in ('1', 'true', 'sim', 'yes', 'on')
    mes_inicio = normalize_month_reference(
        request.args.get('mes') or request.args.get('month') or ('' if todos else mes_atual_real)
    )
    mes_fim = normalize_month_reference(request.args.get('mes_fim') or request.args.get('ate') or '')

    rows = pagamentos_query_rows(mes_inicio=mes_inicio, mes_fim=mes_fim, todos=todos)
    total_month, total_paid_month, total_pending_month, total_gasto, total_investimento, total_terceiros, total_a_pagar = pagamentos_totais_from_rows(rows)

    if todos or not mes_inicio:
        periodo_label = 'Todos'
    elif mes_fim:
        periodo_label = f"{mes_inicio} até {mes_fim}"
    else:
        periodo_label = mes_inicio

    return render_template(
        'pagamentos.html',
        rows=rows,
        total=total_month,
        total_pago=total_paid_month,
        total_pendente=total_pending_month,
        total_gasto=total_gasto,
        total_investimento=total_investimento,
        total_terceiros=total_terceiros,
        total_a_pagar=total_a_pagar,
        mes_inicio=mes_inicio,
        mes_fim=mes_fim,
        mes_atual=mes_inicio or mes_atual_real,
        mes_atual_real=mes_atual_real,
        periodo_label=periodo_label,
        rows_mes_atual=rows,
    )



@require_permission('edit_pagamentos')
def pagamentos_save():
    rid = _safe_int_id(request.form.get('id'))
    try:
        saved_id = save_pagamento(request.form, request.files, rid)
    except Exception as exc:
        flash(f'Não foi possível salvar o pagamento: {exc}', 'danger')
        return redirect(url_for('pagamentos'))

    # ── Pagamentos recorrentes ────────────────────────────────
    repetir_meses = _safe_int_id(request.form.get('repetir_meses') or 0)
    modo_valor = (request.form.get('modo_valor') or 'duplicar').strip().lower()
    a_partir = (request.form.get('repetir_a_partir') or 'atual').strip().lower()
    mes_inicio_custom = (request.form.get('repetir_mes_inicio') or '').strip()

    if repetir_meses and repetir_meses > 1:
        # Determina o mês base conforme opção escolhida
        if a_partir == 'custom' and mes_inicio_custom:
            mes_base = _payment_month_or_current(mes_inicio_custom)
        elif a_partir == 'proximo':
            # Próximo mês após o mês do pagamento
            mes_pag = _payment_month_or_current(request.form.get('pagamento_mes') or '')
            try:
                mn, my = int(mes_pag[:2]), int(mes_pag[3:])
            except Exception:
                mn, my = br_now().month, br_now().year
            mn += 1
            if mn > 12:
                mn = 1
                my += 1
            mes_base = f'{mn:02d}/{my}'
        else:
            # 'atual' — começa do mês do pagamento salvo
            mes_base = _payment_month_or_current(request.form.get('pagamento_mes') or '')

        try:
            m_num, m_ano = int(mes_base[:2]), int(mes_base[3:])
        except Exception:
            m_num, m_ano = br_now().month, br_now().year

        # Valor por mês
        valor_original = parse_num(request.form.get('valor') or 0)
        if modo_valor == 'dividir' and repetir_meses > 0:
            valor_por_mes = round(valor_original / repetir_meses, 2)
            # Atualiza o primeiro lançamento com valor dividido
            if valor_por_mes != valor_original and saved_id:
                try:
                    execute('UPDATE pagamentos SET valor=? WHERE id=?',
                            (str(valor_por_mes).replace('.', ','), saved_id))
                except Exception:
                    pass
            # Se a_partir != 'atual', o primeiro salvo já é o mês base — ajusta o mês
            if a_partir != 'atual' and saved_id:
                try:
                    execute('UPDATE pagamentos SET pagamento_mes=? WHERE id=?', (mes_base, saved_id))
                except Exception:
                    pass
        else:
            valor_por_mes = valor_original

        # Quantos lançamentos adicionais criar
        # Se 'atual': o 1º já foi salvo, cria N-1 extras
        # Se 'proximo' ou 'custom': o 1º foi salvo no mês do pagamento, cria N meses extras a partir do mes_base
        n_extras = repetir_meses - 1 if a_partir == 'atual' else repetir_meses

        criados = 0
        cur_m, cur_a = m_num, m_ano
        if a_partir == 'atual':
            # Avança 1 pois o primeiro já foi salvo neste mês
            cur_m += 1
            if cur_m > 12:
                cur_m = 1
                cur_a += 1

        for i in range(n_extras):
            mes_prox = f'{cur_m:02d}/{cur_a}'
            data_prox = {k: v[0] if isinstance(v, list) else v for k, v in dict(request.form).items()}
            data_prox['pagamento_mes'] = mes_prox
            data_prox['id'] = ''
            data_prox['valor'] = str(valor_por_mes).replace('.', ',')
            try:
                save_pagamento(data_prox, None, None)
                criados += 1
            except Exception as exc_rec:
                app.logger.warning('Pagamento recorrente mês %s falhou: %s', mes_prox, exc_rec)
            cur_m += 1
            if cur_m > 12:
                cur_m = 1
                cur_a += 1

        if criados:
            modo_label = f'dividido em {repetir_meses}x de R$ {valor_por_mes:,.2f}'.replace(',','X').replace('.',',').replace('X','.') if modo_valor == 'dividir' else f'R$ {valor_por_mes:,.2f} cada'.replace(',','X').replace('.',',').replace('X','.')
            flash(f'{criados + (1 if a_partir=="atual" else 0)} lançamentos criados — {modo_label}.', 'success')

    backup_company_data(current_company_id())
    clear_view_cache()
    mes_salvo = _payment_month_or_current(request.form.get('pagamento_mes') or '')
    if not (repetir_meses and repetir_meses > 1):
        flash(f'Pagamento salvo.', 'success')
    return redirect(url_for('pagamentos', mes=mes_salvo))



@require_permission('edit_pagamentos')
def pagamentos_parcelar(rid):
    """Parcela/replica um pagamento já existente.
    Atualiza o registro atual (vira parcela 1) e cria os meses seguintes.
    Espera JSON: { meses, modo_valor, a_partir, mes_inicio }
    """
    if not owned_by_current_company('pagamentos', rid):
        return jsonify({'ok': False, 'error': 'Pagamento não encontrado.'}), 404

    data = request.get_json(silent=True) or {}
    meses      = max(2, min(60, int(data.get('meses') or 2)))
    modo_valor = (data.get('modo_valor') or 'duplicar').strip().lower()
    a_partir   = (data.get('a_partir') or 'atual').strip().lower()
    mes_inicio = (data.get('mes_inicio') or '').strip()

    row = row_to_dict(query_one('SELECT * FROM pagamentos WHERE id=?', (rid,)))
    if not row:
        return jsonify({'ok': False, 'error': 'Pagamento não encontrado.'}), 404

    valor_original = parse_num(row.get('valor') or 0)
    valor_por_mes  = round(valor_original / meses, 2) if modo_valor == 'dividir' else valor_original

    # Determina mês base
    mes_pag = _payment_month_or_current(row.get('pagamento_mes') or '')
    try:
        m_num, m_ano = int(mes_pag[:2]), int(mes_pag[3:])
    except Exception:
        agora = br_now()
        m_num, m_ano = agora.month, agora.year

    if a_partir == 'custom' and mes_inicio:
        mb = _payment_month_or_current(mes_inicio)
        try:
            m_num, m_ano = int(mb[:2]), int(mb[3:])
        except Exception:
            pass
    elif a_partir == 'proximo':
        m_num += 1
        if m_num > 12:
            m_num, m_ano = 1, m_ano + 1

    mes_base = f'{m_num:02d}/{m_ano}'

    # Atualiza o registro atual: vira parcela 1
    try:
        execute('UPDATE pagamentos SET valor=?, pagamento_mes=? WHERE id=?',
                (str(valor_por_mes).replace('.', ','), mes_base, rid))
    except Exception as exc:
        return jsonify({'ok': False, 'error': f'Erro ao atualizar registro: {exc}'}), 500

    # Cria parcelas 2..N nos meses seguintes
    cur_m, cur_a = m_num + 1, m_ano
    if cur_m > 12:
        cur_m, cur_a = 1, m_ano + 1

    criados = 0
    campos_copia = ['fornecedor','descricao_servico','tipo_lancamento','nf_proposta',
                    'acao','sc_pedido','aprovado','tipo_documento','numero_documento',
                    'fluxo_status','empresa_id']
    for _ in range(meses - 1):
        mes_prox = f'{cur_m:02d}/{cur_a}'
        copia = {k: row.get(k, '') or '' for k in campos_copia}
        copia.update({
            'pagamento_mes':   mes_prox,
            'valor':           str(valor_por_mes).replace('.', ','),
            'status':          'Não',
            'data_vencimento': '',
            'anexos_orcamento': '[]',
            'anexos_nf':        '[]',
            'anexos_boleto':    '[]',
            'sistema':          '',
            'equipamento':      '',
        })
        fields = list(copia.keys())
        try:
            execute(
                f"INSERT INTO pagamentos({','.join(fields)}) VALUES ({','.join('?'*len(fields))})",
                [copia[f] for f in fields]
            )
            criados += 1
        except Exception as exc_p:
            app.logger.warning('Parcela %s falhou: %s', mes_prox, exc_p)
        cur_m += 1
        if cur_m > 12:
            cur_m, cur_a = 1, cur_a + 1

    backup_company_data(current_company_id())
    clear_view_cache()

    modo_label = (
        f'dividido em {meses}x de R$ {valor_por_mes:,.2f}'.replace(',','X').replace('.',',').replace('X','.')
        if modo_valor == 'dividir'
        else f'{meses} cópias de R$ {valor_por_mes:,.2f}'.replace(',','X').replace('.',',').replace('X','.')
    )
    return jsonify({'ok': True, 'criados': criados + 1, 'label': modo_label, 'mes_base': mes_base})

# ═══════════════════════════════════════════════════════════════
# HUB DE PAGAMENTOS — sub-módulos
# ═══════════════════════════════════════════════════════════════



@require_permission('view_pagamentos')
def pagamentos_aprovacoes_recentes():
    """Retorna aprovações recentes (últimos 2 min) para popup de notificação."""
    where_sql, params = company_where('pagamentos')
    agora = br_now()
    # Busca pagamentos aprovados nos últimos 2 minutos via fluxo_status
    rows = query_all(
        f"""SELECT id, fornecedor, valor, fluxo_status, pagamento_mes
            FROM pagamentos{where_sql}
            AND aprovado='Sim'
            AND fluxo_status LIKE 'Aprovado por %'
            ORDER BY id DESC LIMIT 20""",
        tuple(params)
    )
    # Filtra pelo timestamp no session para não repetir
    last_seen = session.get('_apr_last_seen_id', 0)
    novos = []
    max_id = last_seen
    for r in rows:
        rid = r['id'] or 0
        if rid > last_seen:
            novos.append({
                'id': rid,
                'fornecedor': r['fornecedor'] or '',
                'valor': br_money(parse_num(r['valor'] or 0)),
                'mes': r['pagamento_mes'] or '',
                'por': (r['fluxo_status'] or '').replace('Aprovado por ', ''),
            })
            if rid > max_id:
                max_id = rid
    if max_id > last_seen:
        session['_apr_last_seen_id'] = max_id
        session.modified = True
    return jsonify({'ok': True, 'novos': novos[:5]})



@require_permission('view_pagamentos')
def pagamentos_receber():
    where_sql, params = company_and('recebimentos')
    rows_db = query_all(f"SELECT * FROM recebimentos WHERE 1=1{where_sql} ORDER BY id DESC", tuple(params))
    rows = [dict(r) for r in rows_db]
    total_pendente = sum(float(str(r.get('valor','0')).replace('R$','').replace('.','').replace(',','.').strip() or 0) for r in rows if r.get('status','') == 'Pendente')
    total_recebido = sum(float(str(r.get('valor','0')).replace('R$','').replace('.','').replace(',','.').strip() or 0) for r in rows if r.get('status','') == 'Recebido')
    total_geral = total_pendente + total_recebido
    return render_template('pagamentos_receber.html', rows=rows, total_pendente=total_pendente, total_recebido=total_recebido, total_geral=total_geral)



@require_permission('edit_pagamentos')
def pagamentos_receber_save():
    data = request.form
    rid = _safe_int_id(data.get('id'))
    fields = ['cliente','descricao','valor','status','data_vencimento','data_recebimento','mes_referencia','numero_documento','observacoes','empresa_id']
    payload = {k: str(data.get(k, '') or '').strip() for k in fields}
    payload['empresa_id'] = current_company_id()
    vals = [payload[k] for k in fields]
    if rid and owned_by_current_company('recebimentos', rid):
        execute(f"UPDATE recebimentos SET {','.join(f+'=?' for f in fields)} WHERE id=?", vals + [rid])
    else:
        execute(f"INSERT INTO recebimentos({','.join(fields)}) VALUES ({','.join(['?']*len(fields))})", vals)
    clear_view_cache()
    return redirect(url_for('pagamentos_receber'))



@require_permission('delete_pagamentos')
def pagamentos_receber_delete():
    rid = _safe_int_id(request.form.get('id'))
    if rid and owned_by_current_company('recebimentos', rid):
        execute('DELETE FROM recebimentos WHERE id=?', (rid,))
        clear_view_cache()
    return redirect(url_for('pagamentos_receber'))



@require_permission('view_pagamentos')
def pagamentos_receber_api(rid):
    row = query_one('SELECT * FROM recebimentos WHERE id=?', (rid,))
    if not row or not owned_by_current_company('recebimentos', rid):
        return jsonify({}), 404
    return jsonify(dict(row))



@require_permission('view_pagamentos')
def pagamentos_hub():
    """Hub central do módulo de Pagamentos."""
    empresa_id = current_company_id()
    agora = br_now()
    mes_atual = f"{agora.month:02d}/{agora.year:04d}"

    where_sql, params = company_where('pagamentos')

    # KPIs rápidos para os cards
    rows_mes = query_all(
        f"SELECT valor, status, data_vencimento, tipo_lancamento FROM pagamentos{where_sql} AND pagamento_mes=?",
        tuple(list(params) + [mes_atual])
    )
    total_mes = sum(parse_num(r['valor']) for r in rows_mes)
    pendente_mes = sum(parse_num(r['valor']) for r in rows_mes if not payment_status_is_paid(r['status']))

    # Vencimentos próximos (7 dias)
    hoje = agora.strftime('%d/%m/%Y')
    proximos7 = []
    for r in query_all(f"SELECT fornecedor, valor, data_vencimento, status FROM pagamentos{where_sql} AND data_vencimento != '' AND status != 'Sim'", tuple(params)):
        venc = parse_br_date(r['data_vencimento'] or '')
        if venc and 0 <= (venc.date() - agora.date()).days <= 7:
            proximos7.append(r)

    # Aprovações pendentes
    aprovacoes = query_all(
        f"SELECT COUNT(*) AS c FROM pagamentos{where_sql} AND (aprovado IS NULL OR aprovado = '' OR aprovado = 'Não') AND status != 'Sim'",
        tuple(params)
    )
    pendentes_aprovacao = int((aprovacoes[0]['c'] if aprovacoes else 0))

    # Fornecedores únicos
    forn_rows = query_all(f"SELECT DISTINCT fornecedor FROM pagamentos{where_sql} AND fornecedor != ''", tuple(params))
    total_fornecedores = len(forn_rows)

    return render_template('pagamentos_hub.html',
        total_mes=total_mes,
        pendente_mes=pendente_mes,
        vencimentos_proximos=len(proximos7),
        pendentes_aprovacao=pendentes_aprovacao,
        total_fornecedores=total_fornecedores,
        mes_atual=mes_atual,
    )




@require_permission('view_pagamentos')
def pagamentos_vencimentos():
    """Calendário de vencimentos."""
    agora = br_now()
    mes_ref = request.args.get('mes', f"{agora.month:02d}/{agora.year:04d}")
    try:
        mes_num, ano_num = int(mes_ref[:2]), int(mes_ref[3:])
    except Exception:
        mes_num, ano_num = agora.month, agora.year

    where_sql, params = company_where('pagamentos')
    rows = query_all(
        f"SELECT id, fornecedor, valor, status, data_vencimento, pagamento_mes, tipo_lancamento, descricao_servico FROM pagamentos{where_sql} AND data_vencimento != ''",
        tuple(params)
    )

    # Agrupa por dia do mês
    from collections import defaultdict
    import calendar
    dias = defaultdict(list)
    totais_dia = defaultdict(float)
    pagos_dia = defaultdict(float)

    for r in rows:
        venc = parse_br_date(r['data_vencimento'] or '')
        if not venc or venc.month != mes_num or venc.year != ano_num:
            continue
        dia = venc.day
        valor = parse_num(r['valor'])
        # Calcula urgência do registro
        dias_diff = (venc.date() - agora.date()).days
        if payment_status_is_paid(r['status']):
            urgencia = 'pago'
        elif dias_diff < 0:
            urgencia = 'vencido'
        elif dias_diff == 0:
            urgencia = 'hoje'
        elif dias_diff <= 3:
            urgencia = 'proximo'
        else:
            urgencia = 'ok'
        r_dict = dict(r)
        r_dict['valor_num'] = valor
        r_dict['urgencia'] = urgencia
        dias[dia].append(r_dict)
        totais_dia[dia] += valor
        if payment_status_is_paid(r['status']):
            pagos_dia[dia] += valor

    # Calendário do mês
    cal = calendar.monthcalendar(ano_num, mes_num)
    hoje_dia = agora.day if agora.month == mes_num and agora.year == ano_num else None
    meses_pt = {1:'Janeiro',2:'Fevereiro',3:'Março',4:'Abril',5:'Maio',6:'Junho',
                7:'Julho',8:'Agosto',9:'Setembro',10:'Outubro',11:'Novembro',12:'Dezembro'}

    # Mês anterior e próximo
    mes_ant = f"{mes_num-1:02d}/{ano_num}" if mes_num > 1 else f"12/{ano_num-1}"
    mes_prox = f"{mes_num+1:02d}/{ano_num}" if mes_num < 12 else f"01/{ano_num+1}"

    # Urgência por dia (pior urgência do dia)
    _urgprio = {'vencido':0,'hoje':1,'proximo':2,'ok':3,'pago':4}
    urgencia_dia = {}
    for dia, itens in dias.items():
        pendentes_dia = [i for i in itens if i.get('urgencia') != 'pago']
        if pendentes_dia:
            urgencia_dia[dia] = min(pendentes_dia, key=lambda x: _urgprio.get(x.get('urgencia','ok'),3))['urgencia']
        else:
            urgencia_dia[dia] = 'pago'

    return render_template('pagamentos_vencimentos.html',
        cal=cal,
        dias=dict(dias),
        totais_dia=dict(totais_dia),
        pagos_dia=dict(pagos_dia),
        urgencia_dia=urgencia_dia,
        mes_ref=mes_ref,
        mes_num=mes_num,
        ano_num=ano_num,
        mes_nome=meses_pt.get(mes_num, ''),
        hoje_dia=hoje_dia,
        mes_ant=mes_ant,
        mes_prox=mes_prox,
        total_mes=sum(totais_dia.values()),
        total_pago=sum(pagos_dia.values()),
        total_pendente=sum(totais_dia.values()) - sum(pagos_dia.values()),
    )




@require_permission('view_pagamentos')
def pagamentos_fornecedores():
    """Perfis de fornecedores gerados automaticamente dos lançamentos."""
    where_sql, params = company_where('pagamentos')
    q = (request.args.get('q') or '').strip()

    rows = query_all(
        f"SELECT fornecedor, valor, status, pagamento_mes, tipo_lancamento, data_vencimento, descricao_servico FROM pagamentos{where_sql} AND fornecedor != ''",
        tuple(params)
    )

    # Agrupa por fornecedor
    from collections import defaultdict
    fornecedores = defaultdict(lambda: {
        'total': 0, 'pago': 0, 'pendente': 0, 'count': 0,
        'meses': set(), 'tipos': set(), 'ultimo_mes': '',
        'descricoes': [], 'vencimentos': []
    })

    for r in rows:
        nome = (r['fornecedor'] or '').strip()
        if not nome:
            continue
        val = parse_num(r['valor'])
        f = fornecedores[nome]
        f['total'] += val
        f['count'] += 1
        if payment_status_is_paid(r['status']):
            f['pago'] += val
        else:
            f['pendente'] += val
        if r['pagamento_mes']:
            f['meses'].add(r['pagamento_mes'])
            if not f['ultimo_mes'] or r['pagamento_mes'] > f['ultimo_mes']:
                f['ultimo_mes'] = r['pagamento_mes']
        if r['tipo_lancamento']:
            f['tipos'].add(r['tipo_lancamento'])
        desc = (r['descricao_servico'] or '').strip()
        if desc and desc not in f['descricoes']:
            f['descricoes'].append(desc)
        if r['data_vencimento']:
            f['vencimentos'].append(r['data_vencimento'])

    # Converte para lista e ordena por total
    lista = []
    for nome, f in fornecedores.items():
        if q and q.lower() not in nome.lower():
            continue
        lista.append({
            'nome': nome,
            'total': f['total'],
            'pago': f['pago'],
            'pendente': f['pendente'],
            'count': f['count'],
            'meses': len(f['meses']),
            'ultimo_mes': f['ultimo_mes'],
            'tipos': list(f['tipos']),
            'descricoes': f['descricoes'][:3],
            'iniciais': ''.join(p[0].upper() for p in nome.split()[:2]),
        })
    lista.sort(key=lambda x: x['total'], reverse=True)

    total_geral = sum(f['total'] for f in lista)

    return render_template('pagamentos_fornecedores.html',
        fornecedores=lista,
        total_geral=total_geral,
        q=q,
    )




@require_permission('view_pagamentos')
def pagamentos_aprovacoes():
    """Fila de aprovação de pagamentos."""
    where_sql, params = company_where('pagamentos')
    agora = br_now()

    todos = [row_to_dict(r) for r in query_all(
        f"""SELECT id, fornecedor, descricao_servico, valor, status, pagamento_mes,
                   data_vencimento, tipo_lancamento, aprovado, fluxo_status,
                   numero_documento, sc_pedido, anexos_nf, anexos_boleto, anexos_orcamento
            FROM pagamentos{where_sql}
            AND status != 'Sim'
            ORDER BY data_vencimento ASC, id DESC""",
        tuple(params)
    )]

    hoje = agora.date()
    pendentes = []
    aprovados  = []

    for r in todos:
        val_raw = r.get('valor') or ''
        val_num = parse_num(val_raw)
        r['valor_num'] = val_num
        r['valor_fmt'] = br_money(val_num) if val_num else (val_raw or '—')
        venc = parse_br_date(r.get('data_vencimento') or '')
        if venc:
            dias = (venc.date() - hoje).days
            r['dias_venc'] = dias
            r['urgencia'] = 'vencido' if dias < 0 else ('hoje' if dias == 0 else ('amanha' if dias == 1 else ('semana' if dias <= 7 else 'normal')))
        else:
            r['dias_venc'] = None
            r['urgencia'] = 'normal'
        apr = (r.get('aprovado') or '').strip().lower()
        is_apr = apr in ('sim','s','yes','1','aprovado')
        r['aprovado_label'] = 'Aprovado' if is_apr else ('Reprovado' if apr in ('não','nao','n','no','reprovado') else 'Aguardando')
        r['anexos_count'] = len(r.get('anexos_nf') or []) + len(r.get('anexos_boleto') or []) + len(r.get('anexos_orcamento') or [])
        if is_apr:
            aprovados.append(r)
        else:
            pendentes.append(r)

    total_pendente = sum(r['valor_num'] for r in pendentes)
    total_aprovado = sum(r['valor_num'] for r in aprovados)
    vencidos = [r for r in pendentes if r['urgencia'] == 'vencido']
    urgentes = [r for r in pendentes if r['urgencia'] in ('hoje', 'amanha', 'semana')]

    return render_template('pagamentos_aprovacoes.html',
        pendentes=pendentes,
        aprovados=aprovados,
        total_pendente=total_pendente,
        total_aprovado=total_aprovado,
        count_vencidos=len(vencidos),
        count_urgentes=len(urgentes),
        count_total=len(pendentes),
        count_aprovados=len(aprovados),
    )




@require_permission('edit_pagamentos')
def pagamentos_aprovar(rid):
    """Aprovar ou reprovar um pagamento."""
    acao = (request.form.get('acao') or '').strip().lower()
    motivo = (request.form.get('motivo') or '').strip()
    user = get_current_user() or {}

    if acao == 'aprovar':
        execute('UPDATE pagamentos SET aprovado=?, fluxo_status=? WHERE id=?',
                ('Sim', 'Aprovado por ' + (user.get('nome') or 'usuário'), rid))
        flash('Pagamento aprovado.', 'success')
    elif acao == 'reprovar':
        execute('UPDATE pagamentos SET aprovado=?, fluxo_status=? WHERE id=?',
                ('Não', 'Reprovado: ' + (motivo or 'sem motivo'), rid))
        flash('Pagamento reprovado.', 'warning')

    clear_view_cache()
    return redirect(url_for('pagamentos_aprovacoes'))




@require_permission('edit_pagamentos')
def pagamentos_aprovar_lote():
    """Aprovar ou reprovar múltiplos pagamentos de uma vez."""
    data = request.get_json(silent=True) or {}
    ids = [int(x) for x in (data.get('ids') or []) if str(x).isdigit()]
    acao = (data.get('acao') or '').strip().lower()
    motivo = (data.get('motivo') or '').strip()
    user = get_current_user() or {}

    if not ids or acao not in ('aprovar', 'reprovar'):
        return jsonify({'ok': False, 'error': 'Parâmetros inválidos.'}), 400

    placeholders = ','.join('?' * len(ids))
    empresa_id = current_company_id()
    # Filtra por IDs + empresa para segurança
    if empresa_id:
        where = f'WHERE id IN ({placeholders}) AND empresa_id=?'
        all_params = ids + [empresa_id]
    else:
        where = f'WHERE id IN ({placeholders})'
        all_params = ids

    if acao == 'aprovar':
        label = 'Aprovado por ' + (user.get('nome') or 'usuário')
        execute(f"UPDATE pagamentos SET aprovado='Sim', fluxo_status=? {where}",
                [label] + all_params)
    else:
        label = 'Reprovado: ' + (motivo or 'sem motivo')
        execute(f"UPDATE pagamentos SET aprovado='Não', fluxo_status=? {where}",
                [label] + all_params)

    clear_view_cache()
    return jsonify({'ok': True, 'updated': len(ids), 'acao': acao})




@require_permission('view_pagamentos')
def pagamentos_relatorios():
    """Relatórios analíticos de pagamentos."""
    where_sql, params = company_where('pagamentos')
    agora = br_now()
    hoje = agora.date()

    rows = [row_to_dict(r) for r in query_all(
        f"SELECT id, fornecedor, valor, status, pagamento_mes, data_vencimento, tipo_lancamento, descricao_servico FROM pagamentos{where_sql}",
        tuple(params)
    )]

    # ── AGING ─────────────────────────────────────────────────
    aging = {'corrente': 0, 'ate30': 0, 'ate60': 0, 'ate90': 0, 'acima90': 0}
    aging_val = {'corrente': 0.0, 'ate30': 0.0, 'ate60': 0.0, 'ate90': 0.0, 'acima90': 0.0}
    for r in rows:
        if payment_status_is_paid(r.get('status')):
            continue
        venc = parse_br_date(r.get('data_vencimento') or '')
        val = parse_num(r.get('valor', 0))
        if not venc:
            aging['corrente'] += 1
            aging_val['corrente'] += val
            continue
        dias = (hoje - venc.date()).days
        if dias <= 0:
            aging['corrente'] += 1; aging_val['corrente'] += val
        elif dias <= 30:
            aging['ate30'] += 1; aging_val['ate30'] += val
        elif dias <= 60:
            aging['ate60'] += 1; aging_val['ate60'] += val
        elif dias <= 90:
            aging['ate90'] += 1; aging_val['ate90'] += val
        else:
            aging['acima90'] += 1; aging_val['acima90'] += val

    # ── CONCENTRAÇÃO POR FORNECEDOR ───────────────────────────
    forn_total = {}
    for r in rows:
        nome = (r.get('fornecedor') or 'Sem fornecedor').strip()
        forn_total[nome] = forn_total.get(nome, 0) + parse_num(r.get('valor', 0))
    total_geral = sum(forn_total.values()) or 1
    concentracao = sorted(
        [{'nome': k, 'total': v, 'perc': round(v / total_geral * 100, 1)} for k, v in forn_total.items()],
        key=lambda x: x['total'], reverse=True
    )[:10]

    # ── TENDÊNCIA 6 MESES ─────────────────────────────────────
    meses_labels = []
    meses_gastos = []
    meses_pagos = []
    for i in range(5, -1, -1):
        m = agora.month - i
        y = agora.year
        while m <= 0:
            m += 12; y -= 1
        ref = f"{m:02d}/{y}"
        meses_labels.append(ref)
        mes_rows = [r for r in rows if r.get('pagamento_mes') == ref]
        meses_gastos.append(round(sum(parse_num(r.get('valor', 0)) for r in mes_rows), 2))
        meses_pagos.append(round(sum(parse_num(r.get('valor', 0)) for r in mes_rows if payment_status_is_paid(r.get('status'))), 2))

    # ── ANOMALIAS ─────────────────────────────────────────────
    from collections import defaultdict
    forn_historico = defaultdict(list)
    for r in rows:
        nome = (r.get('fornecedor') or '').strip()
        val = parse_num(r.get('valor', 0))
        if nome and val > 0:
            forn_historico[nome].append(val)

    anomalias = []
    for nome, valores in forn_historico.items():
        if len(valores) < 3:
            continue
        media = sum(valores) / len(valores)
        ultimo = valores[-1]
        if media > 0 and abs(ultimo - media) / media > 0.5:
            anomalias.append({
                'fornecedor': nome,
                'media': media,
                'ultimo': ultimo,
                'variacao': round((ultimo - media) / media * 100, 1),
            })
    anomalias.sort(key=lambda x: abs(x['variacao']), reverse=True)

    return render_template('pagamentos_relatorios.html',
        aging=aging,
        aging_val=aging_val,
        concentracao=concentracao,
        total_geral=total_geral,
        meses_labels=meses_labels,
        meses_gastos=meses_gastos,
        meses_pagos=meses_pagos,
        anomalias=anomalias[:5],
    )




@require_permission('generate_pdf')
def pagamentos_relatorios_pdf():
    """Gera PDF analítico com aging, concentração e anomalias."""
    from collections import defaultdict
    from reportlab.platypus import HRFlowable
    where_sql, params = company_where('pagamentos')
    agora = br_now()
    hoje = agora.date()
    rows = [row_to_dict(r) for r in query_all(
        f"SELECT id, fornecedor, valor, status, pagamento_mes, data_vencimento, tipo_lancamento FROM pagamentos{where_sql}",
        tuple(params)
    )]
    # Aging
    aging_labels = ['Corrente / Sem venc.','Vencido 1-30d','Vencido 31-60d','Vencido 61-90d','Vencido 90+d']
    aging_keys   = ['corrente','ate30','ate60','ate90','acima90']
    aging     = {k:0   for k in aging_keys}
    aging_val = {k:0.0 for k in aging_keys}
    for r in rows:
        if payment_status_is_paid(r.get('status')): continue
        venc = parse_br_date(r.get('data_vencimento') or '')
        val  = parse_num(r.get('valor', 0))
        if not venc:
            aging['corrente']+=1; aging_val['corrente']+=val; continue
        d = (hoje - venc.date()).days
        k = 'corrente' if d<=0 else ('ate30' if d<=30 else ('ate60' if d<=60 else ('ate90' if d<=90 else 'acima90')))
        aging[k]+=1; aging_val[k]+=val
    # Concentracao
    forn_total = {}
    for r in rows:
        n = (r.get('fornecedor') or 'Sem fornecedor').strip()
        forn_total[n] = forn_total.get(n,0) + parse_num(r.get('valor',0))
    total_geral = sum(forn_total.values()) or 1
    concentracao = sorted(
        [{'nome':k,'total':v,'perc':round(v/total_geral*100,1)} for k,v in forn_total.items()],
        key=lambda x:x['total'], reverse=True
    )[:10]
    # Anomalias
    forn_hist = defaultdict(list)
    for r in rows:
        n=(r.get('fornecedor') or '').strip(); v=parse_num(r.get('valor',0))
        if n and v>0: forn_hist[n].append(v)
    anomalias=[]
    for n,vals in forn_hist.items():
        if len(vals)<3: continue
        media=sum(vals)/len(vals); ultimo=vals[-1]
        if media>0 and abs(ultimo-media)/media>0.5:
            anomalias.append({'fornecedor':n,'media':media,'ultimo':ultimo,'variacao':round((ultimo-media)/media*100,1)})
    anomalias.sort(key=lambda x:abs(x['variacao']),reverse=True)
    # PDF
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=14*mm, rightMargin=14*mm, topMargin=28*mm, bottomMargin=14*mm)
    stl = getSampleStyleSheet()
    s_title = ParagraphStyle('pt', parent=stl['Heading1'], fontSize=14, spaceAfter=2*mm, textColor=colors.HexColor('#111827'))
    s_sec   = ParagraphStyle('ps', parent=stl['Heading2'], fontSize=10, spaceBefore=5*mm, spaceAfter=2*mm, textColor=colors.HexColor('#1d4ed8'))
    s_sub   = ParagraphStyle('pb', parent=stl['Normal'], fontSize=8, textColor=colors.HexColor('#6b7280'))
    s_norm  = ParagraphStyle('pn', parent=stl['Normal'], fontSize=8)
    CH=colors.HexColor('#f1f5f9'); CG=colors.HexColor('#e5e7eb')
    CR=colors.HexColor('#991b1b'); CB=colors.HexColor('#1d4ed8')
    def mk_tbl(data, widths, extra=None):
        t = Table(data, colWidths=widths, repeatRows=1)
        s = [('GRID',(0,0),(-1,-1),.3,CG),('BACKGROUND',(0,0),(-1,0),CH),
             ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8),
             ('LEADING',(0,0),(-1,-1),10),('TOPPADDING',(0,0),(-1,-1),5),
             ('BOTTOMPADDING',(0,0),(-1,-1),5),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
             ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f9fafb')])]
        if extra: s.extend(extra)
        t.setStyle(TableStyle(s)); return t
    e = [
        Paragraph('Relatorio Analitico de Pagamentos', s_title),
        Paragraph(f'Gerado em {agora.strftime("%d/%m/%Y as %H:%M")} - Total: {br_money(total_geral)}', s_sub),
        Spacer(1,4*mm), HRFlowable(width='100%',thickness=.5,color=CG), Spacer(1,4*mm),
        Paragraph('Aging de Contas a Pagar', s_sec),
        mk_tbl([['Faixa','Lancamentos','Valor']]+[[l,str(aging[k]),br_money(aging_val[k])] for l,k in zip(aging_labels,aging_keys)],
               [100*mm,35*mm,45*mm]),
        Spacer(1,4*mm),
        Paragraph('Concentracao por Fornecedor - Top 10', s_sec),
        mk_tbl([['#','Fornecedor','Total','%']]+[[str(i+1),f['nome'],br_money(f['total']),f"{f['perc']}%"] for i,f in enumerate(concentracao)],
               [10*mm,100*mm,45*mm,25*mm]),
        Spacer(1,4*mm),
        Paragraph('Anomalias de Valor Detectadas', s_sec),
    ]
    if anomalias:
        anom_data=[['Fornecedor','Media','Ultimo','Variacao']]
        extra=[]
        for i,a in enumerate(anomalias[:10],1):
            v=f"+{a['variacao']}%" if a['variacao']>0 else f"{a['variacao']}%"
            anom_data.append([a['fornecedor'],br_money(a['media']),br_money(a['ultimo']),v])
            c=CR if a['variacao']>0 else CB
            extra+=[('TEXTCOLOR',(3,i),(3,i),c),('FONTNAME',(3,i),(3,i),'Helvetica-Bold')]
        e.append(mk_tbl(anom_data,[80*mm,40*mm,40*mm,22*mm],extra))
    else:
        e.append(Paragraph('Nenhuma anomalia detectada.', s_norm))
    doc.build(e,
        onFirstPage=lambda c,d:_draw_pdf_header(c,d,'Relatorio de Pagamentos'),
        onLaterPages=lambda c,d:_draw_pdf_header(c,d,'Relatorio de Pagamentos'))
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'relatorio_pagamentos_{agora.strftime("%Y%m%d")}.pdf')



@require_permission('view_pagamentos')
def pagamentos_attachment(rid, grupo, idx):
    key = ATTACHMENT_GROUPS.get((grupo or '').strip().lower())
    if not key:
        return missing_attachment_response('grupo-invalido')
    row = sync_payment_attachments(rid, persist_db=True)
    if not row or not owned_by_current_company('pagamentos', rid):
        return missing_attachment_response('pagamento')
    anexos = list((row or {}).get(key) or [])
    if idx < 0 or idx >= len(anexos):
        return missing_attachment_response('indice')
    stored = anexos[idx]
    return storage_or_local_response(
    stored,
    as_attachment=False,
    kind='pagamentos',
    empresa_id=row.get('empresa_id') or current_company_id()
)



@require_permission('generate_pdf')
def pagamentos_pdf():
    ids_raw = request.args.get('ids') or ''
    ids = [int(x) for x in re.findall(r'\d+', ids_raw)]
    todos = str(request.args.get('todos') or '').strip().lower() in ('1', 'true', 'sim', 'yes', 'on') or bool(ids)
    mes_inicio = normalize_month_reference(request.args.get('mes') or request.args.get('month') or '')
    mes_fim = normalize_month_reference(request.args.get('mes_fim') or request.args.get('ate') or '')
    rows = pagamentos_query_rows(mes_inicio=mes_inicio, mes_fim=mes_fim, todos=todos, ids=ids)
    headers = ['ID','Fornecedor','Descrição','Valor','Pago','SC','Pedido']
    data = [(r.get('id',''), r.get('fornecedor',''), r.get('descricao_servico',''), br_money(r.get('valor_num', r.get('valor'))), r.get('status') or 'Não', r.get('numero_documento',''), r.get('sc_pedido','')) for r in rows]
    titulo = 'Pagamentos selecionados' if ids else 'Pagamentos'
    return send_file(table_pdf(titulo, headers, data), mimetype='application/pdf', as_attachment=True, download_name='pagamentos_selecionados.pdf' if ids else 'pagamentos.pdf')



@require_permission('generate_excel')
def pagamentos_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    todos = str(request.args.get('todos') or '').strip().lower() in ('1', 'true', 'sim', 'yes', 'on')
    mes_inicio = normalize_month_reference(request.args.get('mes') or request.args.get('month') or '')
    mes_fim = normalize_month_reference(request.args.get('mes_fim') or request.args.get('ate') or '')
    rows = pagamentos_query_rows(mes_inicio=mes_inicio, mes_fim=mes_fim, todos=todos or not mes_inicio)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pagamentos'

    # ── Paleta ──────────────────────────────────────────────────
    COR_HEADER   = 'FF1D4ED8'  # azul primário
    COR_GASTO    = 'FFF3F4F6'  # cinza claro
    COR_INVEST   = 'FFDBEAFE'  # azul suave
    COR_TERCEIRO = 'FFDCFCE7'  # verde suave
    COR_PAGO     = 'FFF0FDF4'  # verde muito suave
    COR_TITLE    = 'FF111827'  # texto escuro

    def _fill(hex6): return PatternFill('solid', fgColor=hex6)
    def _font(bold=False, color='FF111827', size=10):
        return Font(bold=bold, color=color, size=size, name='Calibri')
    def _border():
        s = Side(style='thin', color='FFE5E7EB')
        return Border(left=s, right=s, top=s, bottom=s)
    def _center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def _left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)
    def _right():  return Alignment(horizontal='right',  vertical='center')

    # ── Título do relatório ──────────────────────────────────────
    periodo = mes_inicio or 'Todos'
    if mes_fim: periodo += f' até {mes_fim}'
    ws.merge_cells('A1:L1')
    ws['A1'] = f'Relatório de Pagamentos — {periodo}'
    ws['A1'].font = Font(bold=True, color='FFFFFFFF', size=13, name='Calibri')
    ws['A1'].fill = _fill(COR_HEADER)
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 28

    # ── Resumo por tipo ──────────────────────────────────────────
    _, _, _, tot_gasto, tot_invest, tot_terceiros, tot_a_pagar = pagamentos_totais_from_rows(rows)
    resumo = [
        ('Gasto', br_money(tot_gasto), COR_GASTO),
        ('Investimento', br_money(tot_invest), COR_INVEST),
        ('Terceiros', br_money(tot_terceiros), COR_TERCEIRO),
        ('A pagar', br_money(tot_a_pagar), 'FFFFF7ED'),
    ]
    for i, (label, valor, cor) in enumerate(resumo):
        col = i * 3 + 1
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        cell_label = ws.cell(row=2, column=col, value=label)
        cell_label.font = _font(bold=True, size=9, color='FF6B7280')
        cell_label.fill = _fill(cor)
        cell_label.alignment = _center()
        cell_val = ws.cell(row=3, column=col, value=valor)
        cell_val.font = _font(bold=True, size=11)
        cell_val.fill = _fill(cor)
        cell_val.alignment = _center()
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 22

    # ── Cabeçalhos da tabela ─────────────────────────────────────
    headers = ['#', 'Fornecedor', 'Descrição', 'Tipo', 'Para quem (Terceiro)',
               'Valor', 'Status', 'Vencimento', 'Mês Ref.', 'SC / Doc.', 'Pedido', 'Aprovado']
    col_widths = [5, 28, 40, 14, 24, 16, 12, 14, 10, 12, 12, 12]
    ROW_H = 5

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=ROW_H, column=ci, value=h)
        cell.font = _font(bold=True, color='FFFFFFFF', size=10)
        cell.fill = _fill(COR_HEADER)
        cell.alignment = _center()
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[ROW_H].height = 22

    # ── Dados ────────────────────────────────────────────────────
    tipo_cores = {'Gasto': COR_GASTO, 'Investimento': COR_INVEST, 'Terceiros': COR_TERCEIRO}
    for ri, r in enumerate(rows, ROW_H + 1):
        tipo   = (r.get('tipo_lancamento') or 'Gasto')
        pago   = r.get('status', '') == 'Sim'
        cor_bg = tipo_cores.get(tipo, COR_GASTO)
        if pago: cor_bg = 'FFF0FDF4'
        terceiro_nome = r.get('terceiro_nome') or ''
        valor_num = parse_num(r.get('valor_num', r.get('valor')))

        vals = [
            r.get('id') or '',
            r.get('fornecedor') or '',
            r.get('descricao_servico') or '',
            tipo,
            terceiro_nome if tipo == 'Terceiros' else '',
            valor_num,
            'Pago' if pago else 'Pendente',
            r.get('data_vencimento') or '',
            r.get('pagamento_mes') or '',
            r.get('numero_documento') or '',
            r.get('sc_pedido') or '',
            r.get('aprovado') or '',
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = _fill(cor_bg)
            cell.border = _border()
            cell.font = _font(size=9)
            # Alinhamentos especiais
            if ci == 1:   cell.alignment = _center()
            elif ci == 6: cell.alignment = _right()
            else:         cell.alignment = _left()
            # Formato moeda na coluna valor
            if ci == 6:
                cell.number_format = u'R$ #,##0.00'
        ws.row_dimensions[ri].height = 16

    # ── Freeze panes e filtro ────────────────────────────────────
    ws.freeze_panes = f'A{ROW_H+1}'
    ws.auto_filter.ref = f'A{ROW_H}:{get_column_letter(len(headers))}{ROW_H}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'pagamentos_{(mes_inicio or "todos").replace("/","-")}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)




def register_routes(app):
    rules = [
        ('/pagamentos/import', 'pagamentos_import', pagamentos_import, ['POST']),
        ('/pagamentos', 'pagamentos_redirect', pagamentos_redirect, ['GET']),
        ('/pagamentos/lancamentos', 'pagamentos', pagamentos, ['GET']),
        ('/pagamentos/save', 'pagamentos_save', pagamentos_save, ['POST']),
        ('/api/pagamentos/<int:rid>/parcelar', 'pagamentos_parcelar', pagamentos_parcelar, ['POST']),
        ('/api/pagamentos/aprovacoes-recentes', 'pagamentos_aprovacoes_recentes', pagamentos_aprovacoes_recentes, ['GET']),
        ('/pagamentos/receber', 'pagamentos_receber', pagamentos_receber, ['GET']),
        ('/pagamentos/receber/save', 'pagamentos_receber_save', pagamentos_receber_save, ['POST']),
        ('/pagamentos/receber/delete', 'pagamentos_receber_delete', pagamentos_receber_delete, ['POST']),
        ('/api/pagamentos/receber/<int:rid>', 'pagamentos_receber_api', pagamentos_receber_api, ['GET']),
        ('/pagamentos/hub', 'pagamentos_hub', pagamentos_hub, ['GET']),
        ('/pagamentos/vencimentos', 'pagamentos_vencimentos', pagamentos_vencimentos, ['GET']),
        ('/pagamentos/fornecedores', 'pagamentos_fornecedores', pagamentos_fornecedores, ['GET']),
        ('/pagamentos/aprovacoes', 'pagamentos_aprovacoes', pagamentos_aprovacoes, ['GET']),
        ('/pagamentos/aprovacoes/aprovar/<int:rid>', 'pagamentos_aprovar', pagamentos_aprovar, ['POST']),
        ('/api/pagamentos/aprovacoes/lote', 'pagamentos_aprovar_lote', pagamentos_aprovar_lote, ['POST']),
        ('/pagamentos/relatorios', 'pagamentos_relatorios', pagamentos_relatorios, ['GET']),
        ('/pagamentos/relatorios/pdf', 'pagamentos_relatorios_pdf', pagamentos_relatorios_pdf, ['GET']),
        ('/pagamentos/anexo/<int:rid>/<grupo>/<int:idx>', 'pagamentos_attachment', pagamentos_attachment, ['GET']),
        ('/pagamentos/pdf', 'pagamentos_pdf', pagamentos_pdf, ['GET']),
        ('/pagamentos/excel', 'pagamentos_excel', pagamentos_excel, ['GET']),
    ]
    for rule, endpoint, view, methods in rules:
        app.add_url_rule(rule, endpoint, view, methods=methods)
