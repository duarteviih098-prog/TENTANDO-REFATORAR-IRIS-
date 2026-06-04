"""Utilitários de importação Excel compartilhados entre módulos."""
from datetime import datetime
from app.shared.formatters import br_money, br_now, now_str, parse_num
from app.shared.rows import row_get_value, row_to_dict

from openpyxl import load_workbook

def current_company_id():
    from app.auth import current_company_id as fn
    return fn()


def current_company():
    from app.auth import current_company as fn
    return fn()


def current_user_is_super_admin():
    from app.auth import current_user_is_super_admin as fn
    return fn()


def query_one(sql, params=()):
    from app.db import query_one as db_query_one
    return db_query_one(sql, params)


def query_all(sql, params=()):
    from app.db import query_all as db_query_all
    return db_query_all(sql, params)


def execute(sql, params=()):
    from app.db import execute as db_execute
    return db_execute(sql, params)


def table_columns(table):
    from app.db import table_columns as fn
    return fn(table)


def table_has_column(table, column):
    from app.db import table_has_column as fn
    return fn(table, column)


def select_existing_columns(table, desired, fallback='id'):
    from app.db.schema import select_existing_columns as fn
    return fn(table, desired, fallback=fallback)


def ensure_column(table, column, col_type):
    from app.db import ensure_column as fn
    return fn(table, column, col_type)


def normalize_import_header(text):
    raw = str(text or '').strip().lower()
    raw = raw.replace('ç','c').replace('ã','a').replace('á','a').replace('à','a').replace('â','a').replace('é','e').replace('ê','e').replace('í','i').replace('ó','o').replace('ô','o').replace('õ','o').replace('ú','u')
    raw = re.sub(r'[^a-z0-9]+', '_', raw).strip('_')
    return raw




def excel_rows_from_upload(file_storage):
    wb = load_workbook(file_storage, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [normalize_import_header(v) for v in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if not row or not any(v not in (None, '') for v in row):
            continue
        item = {}
        for idx, head in enumerate(headers):
            if not head:
                continue
            value = row[idx] if idx < len(row) else ''
            if isinstance(value, datetime):
                value = value.strftime('%d/%m/%Y')
            elif value is None:
                value = ''
            else:
                value = str(value).strip()
            item[head] = value
        if item:
            data_rows.append(item)
    return headers, data_rows


